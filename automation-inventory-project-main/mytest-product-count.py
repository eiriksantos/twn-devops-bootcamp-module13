import openpyxl

xl_file = openpyxl.load_workbook("/Users/ericksonsantos/EJ/twn-devops-bootcamp/twn-devops-bootcamp-module13/automation-inventory-project-main/inventory.xlsx")
product_list = xl_file["Sheet1"]
product_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):

    # Accessing and saving the value on each row starting in row 2 column 4 to max row column 4 which is the supplier name
    supplier_name = product_list.cell(product_row, 4).value

    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    
    # Calculation for the total number of products per supplier
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier.update({supplier_name:current_num_products+1})
    else:
        product_per_supplier.update({supplier_name:1})

    
    # Calculation for the total number of value (price x inventory number) per supplier
    if supplier_name in total_value_per_supplier:
        current_value_per_supplier = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier.update({supplier_name:current_value_per_supplier + price * inventory})
    else:
        total_value_per_supplier.update({supplier_name:price * inventory})
    
    #
    if inventory < 10:
        products_under_10_inv.update({f"product {int(product)}":inventory})

    #
    inventory_price.value = inventory * price
    

print(product_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)
xl_file.save("/Users/ericksonsantos/EJ/twn-devops-bootcamp/twn-devops-bootcamp-module13/automation-inventory-project-main/new-inventory.xlsx")